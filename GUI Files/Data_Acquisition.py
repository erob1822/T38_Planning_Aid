"""
Data_Acquisition.py - Unified Data Fetcher for T-38 PlanAid
MERGED VERSION: Combines aengl5337's Cycle Caching with erob1822's Data Processing.

Purpose:
    Acquires all external data needed by the KML generator.
    Uses 'CycleCache' to prevent redundant downloads of large FAA files.
    Deploys active data to the standard DATA/ directories for downstream processing.

Data Sources:
    1. FAA NASR (28-day cycle) -> DATA/apt_data/
    2. FAA DCS (56-day cycle) -> DATA/afd/ & DATA/jasu_data.csv
    3. NASA AOD Flight API -> DATA/flights_data.csv
    4. DLA Fuel -> DATA/fuel_data.csv
    5. Google Sheet Comments -> DATA/comments_data.csv
"""

import os
import csv
import shutil
import zipfile
import json
import logging
import traceback
import time
import re
from datetime import datetime, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import deepcopy

# Third-party imports
import requests
import pandas as pd
from requests.adapters import HTTPAdapter
from urllib3.util import Retry
from requests_ntlm import HttpNtlmAuth
import urllib3
from tqdm import tqdm

# Attempt to import PyMuPDF for JASU parsing
try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

# Suppress SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# UTILITIES & CONFIG
# ---------------------------------------------------------------------------

def create_session():
    """Configure HTTP session with retries (including on timeouts)."""
    retry_strategy = Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "HEAD", "OPTIONS"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    http = requests.Session()
    http.mount("https://", adapter)
    http.mount("http://", adapter)
    return http

HTTP_SESSION = create_session()

# ---------------------------------------------------------------------------
# CORE CLASSES (Derived from aengl5337)
# ---------------------------------------------------------------------------

class DataSource:
    """Represents a single data source (NASR, DCS, Flights, etc.) with state management."""

    def __init__(self, name, config, download_method, cycle_fetch_method=None, deploy_method=None, **kwargs):
        self.name = name
        self.config = config # The global AppConfig
        self.download_method = download_method
        self.cycle_fetch_method = cycle_fetch_method
        self.deploy_method = deploy_method
        
        # State attributes (will be loaded from cache)
        self.success = False
        self.timestamp = None
        self.downloaded_cycle_date = None
        self.current_cycle_date = None
        self.current = False
        self.exists = False
        self.skip_download = False
        self.download_subdir = None
        
        # Merge kwargs (cached state) into self
        for key, value in kwargs.items():
            setattr(self, key, value)

        # Ensure Paths are objects
        if self.download_subdir and isinstance(self.download_subdir, str):
            self.download_subdir = Path(self.download_subdir)
            
        # Define base download folder for this source (within DATA/Cache)
        self.source_cache_folder = self.config.data_folder / "Cache" / self.name
        self.source_cache_folder.mkdir(parents=True, exist_ok=True)

    def get_state_dict(self) -> dict:
        """Return dynamic state for JSON serialization."""
        return {
            'success': self.success,
            'timestamp': self.timestamp,
            'downloaded_cycle_date': self.downloaded_cycle_date,
            'download_subdir': str(self.download_subdir) if self.download_subdir else None
        }

    def check_cycle_status(self):
        """Check if the source needs updating based on cycle dates."""
        if callable(self.cycle_fetch_method):
            # Updates self.current_cycle_date and potentially self.download_url
            self.cycle_fetch_method(self)
        
        # Determine if we have the current version
        if self.current_cycle_date and self.downloaded_cycle_date:
            self.current = (self.current_cycle_date == self.downloaded_cycle_date)
        else:
            self.current = False

    def should_skip_download(self):
        """Determine if we can skip download and use cached data."""
        self.skip_download = False
        self.exists = False
        
        # Check if physical files exist
        if self.download_subdir and self.download_subdir.exists():
            # Check if directory has files
            self.exists = any(f.is_file() for f in self.download_subdir.rglob('*'))
        
        if self.success and self.exists and self.current:
            self.skip_download = True
            logger.info(f"[{self.name}] Cache Hit: Local data ({self.downloaded_cycle_date}) is current. Skipping download.")
        else:
            if not self.current and self.current_cycle_date:
                 logger.info(f"[{self.name}] Update Required: Local ({self.downloaded_cycle_date}) != Remote ({self.current_cycle_date}).")
            elif not self.exists:
                 logger.info(f"[{self.name}] Cache Miss: Data missing from disk.")
            
            # Reset state for fresh download
            self.success = False
            self.downloaded_cycle_date = None

    def execute(self):
        """Orchestrate the download or retrieval of data, followed by deployment."""
        
        # 1. Check Cycle / URL
        self.check_cycle_status()
        
        # 2. Check Cache
        self.should_skip_download()
        
        if not self.skip_download:
            # 3. Download
            logger.info(f"[{self.name}] Starting download...")
            try:
                # Timestamp the download
                self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # Create timestamped subdir or edition subdir
                folder_name = self.current_cycle_date if self.current_cycle_date else self.timestamp
                self.download_subdir = self.source_cache_folder / folder_name
                self.download_subdir.mkdir(parents=True, exist_ok=True)
                
                # Run specific download logic
                self.download_method(self)
                
                # Mark success
                self.success = True
                if self.current_cycle_date:
                    self.downloaded_cycle_date = self.current_cycle_date
                
                logger.info(f"[{self.name}] Download successful.")
                
            except Exception as e:
                logger.error(f"[{self.name}] Download failed: {e}")
                logger.debug(traceback.format_exc())
                return # Stop processing this source
        
        # 4. Deploy (Copy from Cache to Working Dir)
        if self.success and callable(self.deploy_method):
            logger.info(f"[{self.name}] Deploying to active folders...")
            try:
                self.deploy_method(self)
            except Exception as e:
                logger.error(f"[{self.name}] Deployment failed: {e}")
                logger.debug(traceback.format_exc())


class CycleCache:
    """Manages the lifecycle of all data sources."""
    cache_filename = 'data_download_cache.json'

    def __init__(self, cfg):
        self.cfg = cfg
        self.cache_filepath = cfg.data_folder / self.cache_filename
        self.cache_data = self._load_cache()
        self.sources = {}
        
        # Define Source Configuration Mapping
        # Maps source names to their handler functions
        self.registry = {
            'nasr': {
                'download': download_and_extract_nasr,
                'cycle': get_faa_cycle_url_nasr,
                'deploy': deploy_nasr
            },
            'dcs': {
                'download': download_and_extract_dcs,
                'cycle': get_faa_cycle_url_dcs,
                'deploy': deploy_dcs_and_parse_jasu
            },
            'flights': {
                'download': download_flights,
                'cycle': None, # No cycle API, just download fresh
                'deploy': deploy_simple_csv_flights
            },
            'fuel': {
                'download': download_fuel,
                'cycle': None,
                'deploy': deploy_simple_csv_fuel
            },
            'comments': {
                'download': download_comments,
                'cycle': None,
                'deploy': deploy_simple_csv_comments
            }
        }

        self._init_sources()

    def _load_cache(self):
        if self.cache_filepath.exists():
            try:
                with open(self.cache_filepath, 'r') as f:
                    return json.load(f)
            except Exception:
                logger.warning("Cache file corrupted, starting fresh.")
        return {'sources': {}}

    def _init_sources(self):
        for name, handlers in self.registry.items():
            cached_state = self.cache_data.get('sources', {}).get(name, {})
            
            self.sources[name] = DataSource(
                name=name,
                config=self.cfg,
                download_method=handlers['download'],
                cycle_fetch_method=handlers['cycle'],
                deploy_method=handlers['deploy'],
                **cached_state
            )

    def run_all(self):
        """Execute all sources."""
        # Parallelize the "small" downloads/checks? 
        # For now, keeping it sequential for simplicity and logging clarity, 
        # but threaded execution is easy to add if needed.
        
        for name, source in self.sources.items():
            source.execute()
        
        self._save_cache()

    def _save_cache(self):
        # Update cache data from source objects
        for name, source in self.sources.items():
            if 'sources' not in self.cache_data:
                self.cache_data['sources'] = {}
            self.cache_data['sources'][name] = source.get_state_dict()
        
        try:
            with open(self.cache_filepath, 'w') as f:
                json.dump(self.cache_data, f, indent=2)
            logger.info("Cache saved.")
        except Exception as e:
            logger.warning(f"Failed to save cache: {e}")

# ---------------------------------------------------------------------------
# SPECIFIC DOWNLOAD/FETCH/DEPLOY FUNCTIONS
# ---------------------------------------------------------------------------

# --- FLIGHTS ---
def download_flights(source):
    """Download flight data from NASA API."""
    api_url = source.config.aod_flights_api
    years_included = source.config.years_included
    
    logger.info(f"Querying NASA API: {api_url}")
    response = HTTP_SESSION.get(
        api_url, 
        verify=False, 
        auth=HttpNtlmAuth('', ''),
        timeout=60
    )
    response.raise_for_status()
    data = response.json()
    
    # Process Data
    cutoff_date = (datetime.now() - timedelta(days=years_included * 365)).strftime('%Y-%m-%d')
    latest_flights = {}

    for entry in data:
        airport_code = entry.get('Airport', '').strip()
        flight_date = entry.get('FlightDate', '')[:10]
        
        if flight_date < cutoff_date:
            continue
        
        # Parse crew
        abvs = entry.get('ABVs', '')
        if isinstance(abvs, str):
            crew = abvs.strip().split(',')
            front = crew[0].strip() if len(crew) > 0 else ''
            back = crew[1].strip() if len(crew) > 1 else ''
        else:
            front, back = '', ''
            
        # Logic: Keep most recent
        if airport_code not in latest_flights or flight_date > latest_flights[airport_code]['date']:
            latest_flights[airport_code] = {
                'icao': airport_code,
                'date': flight_date,
                'front': front,
                'back': back
            }

    # Write to CSV in Download Subdir
    csv_path = source.download_subdir / "flights_data.csv"
    with open(csv_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["ICAO", "DATE_LANDED", "FRONT_SEAT", "BACK_SEAT"])
        for code, info in sorted(latest_flights.items(), key=lambda x: x[1]['date'], reverse=True):
            writer.writerow([code, info['date'], info['front'], info['back']])

def deploy_simple_csv_flights(source):
    src = source.download_subdir / "flights_data.csv"
    dst = source.config.data_folder / "flights_data.csv"
    shutil.copy2(src, dst)

# --- COMMENTS ---
def download_comments(source):
    """Download Google Sheet as CSV."""
    url = source.config.aod_comments_url
    logger.info("Downloading Comments from Google Sheet...")
    # Skip first 3 rows as per original logic
    df = pd.read_csv(url, header=3)
    
    csv_path = source.download_subdir / "comments_data.csv"
    df.to_csv(csv_path, index=False)

def deploy_simple_csv_comments(source):
    src = source.download_subdir / "comments_data.csv"
    dst = source.config.data_folder / "comments_data.csv"
    shutil.copy2(src, dst)

# --- FUEL ---
def download_fuel(source):
    """Download DLA Fuel Data."""
    check_url = source.config.dla_fuel_check
    dl_url = source.config.dla_fuel_download
    
    # Ping check URL (DLA site can be very slow on first connect)
    logger.info("Contacting DLA fuel site (this can take a minute)...")
    HTTP_SESSION.get(check_url, verify=False, timeout=(30, 120))
    
    # Download
    logger.info("Downloading Fuel Data...")
    response = HTTP_SESSION.get(dl_url, verify=False, timeout=(30, 120))
    response.raise_for_status()
    
    csv_path = source.download_subdir / "fuel_data.csv"
    with open(csv_path, 'wb') as f:
        f.write(response.content)

def deploy_simple_csv_fuel(source):
    src = source.download_subdir / "fuel_data.csv"
    dst = source.config.data_folder / "fuel_data.csv"
    shutil.copy2(src, dst)

# --- NASR (Airports) ---
def get_faa_cycle_url_nasr(source):
    _get_faa_cycle_generic(source, source.config.nasr_file_finder)

def download_and_extract_nasr(source):
    _download_and_extract_zip(source, "NASR")

def deploy_nasr(source):
    """Copy specific NASR CSVs to apt_data_dir."""
    required_files = ['APT_BASE.csv', 'APT_RWY.csv', 'APT_RWY_END.csv']
    dest_dir = source.config.apt_data_dir
    dest_dir.mkdir(parents=True, exist_ok=True)
    
    # Find files in the cache structure (they might be in nested folders)
    # Glob ONCE and build a name→path lookup for O(1) access
    search_root = source.download_subdir / "extracted"
    all_csvs = list(search_root.rglob("*.csv"))
    csv_by_name: dict[str, list[Path]] = {}
    for f in all_csvs:
        csv_by_name.setdefault(f.name.lower(), []).append(f)
    logger.debug(f"[NASR] Found {len(all_csvs)} CSVs in extracted tree")

    for req_file in required_files:
        matches = csv_by_name.get(req_file.lower(), [])
        logger.debug(f"[NASR] Matches for {req_file}: {[str(f) for f in matches]}")
        if matches:
            src = max(matches, key=lambda f: f.stat().st_mtime)
            dst = dest_dir / req_file
            shutil.copy2(src, dst)
            logger.info(f"Deployed {req_file} (from {src})")
        else:
            logger.error(f"Could not find {req_file} in NASR download!")

# --- DCS (Digital Chart Supplement) & JASU ---
def get_faa_cycle_url_dcs(source):
    _get_faa_cycle_generic(source, source.config.dcs_file_finder)

def download_and_extract_dcs(source):
    _download_and_extract_zip(source, "DCS")

def deploy_dcs_and_parse_jasu(source):
    """Copy PDFs to DATA/afd and run JASU parser (or use cached jasu_data.csv)."""
    afd_dir = source.config.data_folder / "afd"
    afd_dir.mkdir(parents=True, exist_ok=True)
    
    cached_jasu = source.download_subdir / "jasu_data.csv"
    jasu_dst = source.config.data_folder / "jasu_data.csv"
    
    # Accept an optional progress callback set by the GUI layer
    parse_progress_cb = getattr(source, '_parse_progress_cb', None)
    
    if cached_jasu.exists():
        # JASU already parsed for this cycle — just copy it
        logger.info("[dcs] JASU cache hit — skipping PDF parsing.")
        shutil.copy2(cached_jasu, jasu_dst)
    else:
        # 1. Copy PDFs
        logger.info("Deploying AFD PDFs...")
        pdf_files = list(source.download_subdir.rglob("*.pdf"))
        for pdf in pdf_files:
            shutil.copy2(pdf, afd_dir / pdf.name)
        
        # 2. Parse JASU
        logger.info("Parsing PDFs for JASU data...")
        parse_jasu(source.config, progress_cb=parse_progress_cb)
        
        # 3. Cache the result alongside the DCS download for next run
        if jasu_dst.exists():
            shutil.copy2(jasu_dst, cached_jasu)
            logger.info("[dcs] Cached jasu_data.csv for future runs.")

# --- HELPER: FAA GENERIC ---
def _get_faa_cycle_generic(source, api_url):
    params = {"edition": "current"}
    headers = {"Accept": "application/json"}
    try:
        r = HTTP_SESSION.get(api_url, params=params, headers=headers, timeout=10)
        r.raise_for_status()
        data = r.json()
        
        source.download_url = data["edition"][0]["product"]["url"]
        raw_date = data["edition"][0]["editionDate"]
        # Format: MM/DD/YYYY -> YYYY-MM-DD
        source.current_cycle_date = datetime.strptime(raw_date, "%m/%d/%Y").strftime("%Y-%m-%d")
        logger.info(f"[{source.name}] Current Cycle: {source.current_cycle_date}")
    except Exception as e:
        logger.error(f"Failed to fetch FAA cycle info: {e}")

def _download_and_extract_zip(source, label):
    zip_path = source.download_subdir / "data.zip"
    
    # Download
    logger.info(f"Downloading {label} zip...")
    with HTTP_SESSION.get(source.download_url, stream=True, timeout=300) as r:
        r.raise_for_status()
        total_size = int(r.headers.get('content-length', 0))
        with open(zip_path, 'wb') as f, tqdm(
            total=total_size, unit='B', unit_scale=True, desc=label
        ) as pbar:
            for chunk in r.iter_content(chunk_size=131072):
                f.write(chunk)
                pbar.update(len(chunk))
    
    # Extract
    logger.info(f"Extracting {label}...")
    extract_dir = source.download_subdir / "extracted"
    
    # Files we actually need - skip unneeded files with long paths
    needed_patterns = ['CSV_Data', 'APT_BASE', 'APT_RWY', 'APT_RWY_END', '.txt']
    
    def is_needed_file(member_name):
        """Check if this file is needed or can be skipped."""
        # Skip AIXM/SAA schema files - they have very long paths and aren't needed
        skip_patterns = ['AIXM', 'SAA-AIXM', 'Schema', '.xsd']
        if any(p in member_name for p in skip_patterns):
            return False
        return True
    
    with zipfile.ZipFile(zip_path, 'r') as zf:
        for member in zf.namelist():
            try:
                # Skip files we don't need (prevents long path issues on Windows)
                if not is_needed_file(member):
                    continue
                zf.extract(member, extract_dir)
            except (OSError, FileNotFoundError, KeyError) as e:
                # Handle Windows long path issues - skip if it's a non-essential file
                if 'coordinateReferenceSystems.xsd' in member or 'AIXM' in member or '.xsd' in member:
                    logger.debug(f"Skipping problematic file: {member}")
                    continue
                logger.error(f"Extraction failed for {member}: {e}")
                raise
    
    # Handle nested zips (common in NASR)
    inner_zips = list(extract_dir.rglob("*.zip"))
    for iz in inner_zips:
        try:
            with zipfile.ZipFile(iz, 'r') as zf:
                # Extract into a subfolder named after the zip
                subdir = iz.parent / iz.stem
                for member in zf.namelist():
                    try:
                        if not is_needed_file(member):
                            continue
                        zf.extract(member, subdir)
                    except (OSError, FileNotFoundError) as e:
                        logger.debug(f"Skipping problematic nested file: {member}")
                        continue
        except Exception as e:
            logger.warning(f"Could not process nested zip {iz.name}: {e}")

# ---------------------------------------------------------------------------
# JASU PARSING LOGIC (From erob1822)
# ---------------------------------------------------------------------------

def parse_jasu(cfg, progress_cb=None):
    """Parse DCS/AFD PDF files for airports with JASU.
    
    Args:
        cfg: AppConfig instance.
        progress_cb: Optional callable(current_index, total, pdf_name) for GUI progress.
    """
    if not HAS_FITZ:
        logger.warning("PyMuPDF (fitz) not installed. Skipping JASU parsing.")
        return

    afd_dir = cfg.data_folder / "afd"
    
    def extract_icao_jasu_from_text(text: str) -> list:
        lines = text.splitlines()
        kval_list = []
        JASU_ref = None
        buzzkill = 1
        
        for idx, line in enumerate(lines):
            match = re.search(r'\(\s*(K|PA)\s*(\w{3,4})\s*\)', line)
            if match:
                kval_list.append(idx)
            
            if 'JASU' in line:
                JASU_ref = idx
                for equipment in ['95', '60A', 'MSU', 'GTC', 'WELLS', 'NCPP', 'MA-']:
                    if equipment in line or (idx + 1 < len(lines) and equipment in lines[idx + 1]):
                        buzzkill = 0
                        break
        
        extracted = []
        if JASU_ref is not None and buzzkill == 0:
            prior_icaos = [i for i in kval_list if i < JASU_ref]
            if prior_icaos:
                corr_port = prior_icaos[-1]
                match = re.search(r'\(\s*(K|PA)\s*(\w{3,4})\s*\)', lines[corr_port])
                if match:
                    extracted.append(f"{match.group(1)}{match.group(2)}")
        return extracted

    def process_pdf(pdf_path: Path):
        icaos = []
        try:
            with fitz.open(str(pdf_path)) as doc:
                for page in doc:
                    icaos.extend(extract_icao_jasu_from_text(page.get_text()))
        except Exception:
            pass
        return icaos

    # Filter PDFs
    pdf_files = list(afd_dir.glob("*.pdf"))
    if not pdf_files:
        logger.warning("No PDFs found for JASU parsing.")
        return

    jasu_airports = set()
    total_pdfs = len(pdf_files)
    logger.info(f"Parsing {total_pdfs} PDFs for JASU data...")
    
    with ThreadPoolExecutor(max_workers=min(8, (os.cpu_count() or 4))) as executor:
        futures = {executor.submit(process_pdf, p): p for p in pdf_files}
        for i, future in enumerate(as_completed(futures), 1):
            pdf_path = futures[future]
            if progress_cb:
                progress_cb(i, total_pdfs, pdf_path.stem)
            jasu_airports.update(future.result())

    # Write Result
    jasu_path = cfg.data_folder / "jasu_data.csv"
    with open(jasu_path, 'w') as f:
        f.write("ICAO\n")
        for icao in sorted(jasu_airports):
            f.write(f"{icao}\n")
    logger.info(f"JASU data written to {jasu_path}")

# ---------------------------------------------------------------------------
# EXCEL UPDATE LOGIC (From erob1822)
# ---------------------------------------------------------------------------

def update_wb_list(cfg):
    """Update wb_list.xlsx with recent data."""
    import openpyxl
    wb_path = cfg.work_dir / 'wb_list.xlsx'
    
    if not wb_path.exists():
        logger.error("wb_list.xlsx not found.")
        return

    try:
        wb = openpyxl.load_workbook(wb_path)
        if 'kml data' not in wb.sheetnames:
            logger.error("'kml data' sheet missing.")
            return
        
        ws = wb['kml data']
        header_row = 1
        headers = {str(cell.value).strip().upper(): cell.column for cell in ws[header_row] if cell.value}

        # Update Flights
        flights_path = cfg.data_folder / "flights_data.csv"
        if flights_path.exists():
            df = pd.read_csv(flights_path)
            cols = ['RECENTLY_LANDED', 'DATE_LANDED', 'FRONT_SEAT', 'BACK_SEAT']
            
            # Clear old
            for h in cols:
                if h in headers:
                    for row in range(header_row+1, ws.max_row+1):
                        ws.cell(row=row, column=headers[h], value=None)
            
            # Write new
            for idx, row in df.iterrows():
                for h in cols:
                    if h in headers and h in row:
                        ws.cell(row=header_row+1+idx, column=headers[h], value=row[h])

        # Update Comments
        comm_path = cfg.data_folder / "comments_data.csv"
        if comm_path.exists():
            df = pd.read_csv(comm_path)
            cols = ['APT_COMM', 'COMMENT_DATE', 'COMMENTS']
            
            # Clear old
            for h in cols:
                if h in headers:
                    for row in range(header_row+1, ws.max_row+1):
                        ws.cell(row=row, column=headers[h], value=None)
            
            # Write new
            for idx, row in df.iterrows():
                for h in cols:
                    if h in headers and h in row:
                        ws.cell(row=header_row+1+idx, column=headers[h], value=row[h])

        wb.save(wb_path)
        logger.info("wb_list.xlsx updated successfully.")

    except PermissionError:
        logger.error("wb_list.xlsx is locked by another process (close Excel?).")
        raise  # Let callers detect the file-lock and offer a retry
    except Exception as e:
        logger.error(f"Failed to update wb_list.xlsx: {e}")
        raise  # Let callers detect the failure

# ---------------------------------------------------------------------------
# MAIN ENTRY POINT
# ---------------------------------------------------------------------------

def run(cfg):
    """Main execution function called by T38_PlanAid.py."""
    
    # 1. Run Cycle Cache Manager (Downloads & Deploys)
    cache_mgr = CycleCache(cfg)
    cache_mgr.run_all()
    
    # 3. Update User Excel File
    update_wb_list(cfg)
    
    logger.info("Data Acquisition Complete.")

if __name__ == "__main__":
    print("This module is intended to be imported by T38_PlanAid.py")